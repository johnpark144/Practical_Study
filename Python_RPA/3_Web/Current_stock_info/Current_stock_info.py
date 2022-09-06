from openpyxl.styles import Font, Alignment, PatternFill, Color
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
options = webdriver.ChromeOptions()
options.headless = True # To execute it without pop-up
options.add_argument('--log-level=1') # To accesss thru headless
driver = webdriver.Chrome("./chromedriver.exe",options=options)

def stock_movers():
    def data(url, color):
        driver.get(url)
        time.sleep(2)
        ws.sheet_properties.tabColor = color # To differentiate sheet by sheet
        ws.column_dimensions["B"].width = 35 # enlarge the column, due to name_size
        ws.column_dimensions["E"].width = 13
        ws.row_dimensions.height = 20
        ws.row_dimensions[1].height = 25

        ws["A1"] = "Symbol"
        ws["B1"] = "Company Name"
        ws["C1"] = "Price"
        ws["D1"] = "Change"
        ws["E1"] = "% Change"
        ws["F1"] = "Volume"

        for idx in range(1,7) : # 1st row decoration
            ws.cell(column= idx, row= 1).font = Font(bold=True, size=12)
            ws.cell(column= idx, row= 1).alignment = Alignment(horizontal='center')
            ws.cell(column= idx, row= 1).fill = PatternFill(fill_type='solid', fgColor=Color(color))
            
        for idx in range(1,101) :
            row = driver.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]'.format(idx)) # One row
            ws.cell(column= 1, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[1]'.format(idx)).text) # Symbol
            ws.cell(column= 2, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[2]'.format(idx)).text) # Company Name
            ws.cell(column= 3, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[3]'.format(idx)).text) # Price
            ws.cell(column= 4, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[4]'.format(idx)).text) # Change
            ws.cell(column= 5, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[5]'.format(idx)).text) # % Change
            ws.cell(column= 6, row= idx+1, value = row.find_element(By.XPATH,'//*[@id="scr-res-table"]/div[1]/table/tbody/tr[{}]/td[6]'.format(idx)).text) # Volume

    wb = Workbook()
    ws = wb.active
    ws.title = "Most Actives"
    data("https://finance.yahoo.com/most-active?offset=0&count=100", "789abc") # Most active stock link, Bule

    ws = wb.create_sheet("Gainers", 2)
    data("https://finance.yahoo.com/gainers?offset=0&count=100", "9bffda")  # Gainers stock link, Green

    ws = wb.create_sheet("Losers", 3)
    data("https://finance.yahoo.com/losers?offset=0&count=100", "cd3861") # Losers stock link, Red

    driver.close()

    mycttime = time.ctime() # Current time
    file_time = f"{mycttime[4:7]+mycttime[8:10]}{mycttime[11:13]}{mycttime[14:16]}_stock" # Month,day,hour,minuate_stock (When made file)
    print(file_time)

    wb.save(f"{file_time}.xlsx") # Month,day,hour,minuate_stock.xlsx
    
        
if __name__ == "__main__":  # It is not called when called from other modules
    stock_movers()
    print("All current stock datas are saved")
