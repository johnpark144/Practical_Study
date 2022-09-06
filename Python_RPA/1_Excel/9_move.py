from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학
# ws.move_range("B1:C11", rows=0, cols=1) # 지정된 범위를 rows=0만큼 cols=1만큼 움직여라
# ws["B1"].value = "국어" # B1 셀에 '국어'입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=10, cols=-1)

wb.save("sample_korean.xlsx")