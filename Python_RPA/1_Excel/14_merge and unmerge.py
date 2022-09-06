from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2부터 D2까지 합치겠음
ws["B2"].value = "Merged Cell"

# B2:D2 병합되어 있던 셀을 해제
# ws.unmerge_cells("B2:D2")

wb.save("sample_unmerge.xlsx")