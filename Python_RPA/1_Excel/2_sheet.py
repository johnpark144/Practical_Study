from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
wb.active
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws2 = wb.create_sheet("Newsheet", 2) # 2번째 index에 Sheet 생성,(0,1,2,3....)

ws.title = "Mysheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "117F87" # RGB 형태로 값을 넣어주면 탭 색상 변경

new_ws = wb["Newsheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
