from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
ws = wb.active
ws.title = "Nadosheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1셀의 '값'을 출력
print(ws["A10"].value) # 값이 없을 떈 'None' 을 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ... , 위 방식보다 입력하기 어렵지만 반복문을 수월하기엔 쉬움 
print(ws.cell(column=1, row=1).value) # ws["A1"].value
print(ws.cell(column=2, row=1).value) # ws["B1"].value

c = ws.cell(column= 3, row= 1, value=10) # ws["c1"].value = 10
print(c.value) # ws["C1"].value

from random import *

# 반복문을 이용해 랜덤 숫자 채우기
index = 1
for x in range(1, 11) : # 10개 row
    for y in range(1, 11) : # 10개 column
        ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자의 값을 A1~J10셀에 넣기
        ws.cell(row=x+11, column=y, value=index) # index 값을 A12~J22셀에 넣기
        index += 1


wb.save("sample.xlsx")
