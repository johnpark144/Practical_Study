from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # 영어 column 만 가져 오기
# # print(col_B)
# for cell in col_B :
#     print(cell.value)

# row_title = ws["1"] # 1번째 row만 가져오기
# for cell in row_title :
#     print(cell.value)

# col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
# for cols in col_range :
#     for cell in cols :
#         print(cell.value)

# row_range = ws[2:6] # 1번째 줄을 제외하고 2번째 줄에서 6번째 줄까지 가져오기 (일반 슬라이싱과 다르게 6 미만이 아니라 6을 포함)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string
# row_range = ws[2:ws.max_row] # 2번쨰 줄부터 마지막 줄까지
# for rows in row_range :
#     for cell in rows :
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") # A/10, AZ/250
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="") # A
#         print(xy[1], end=" ") # 1
#     print()

# # 전체 rows
# print(tuple(ws.rows))
# print()
# for row in tuple(ws.rows):
#     print(row[1].value)
# for row in ws.iter_rows(): # 위에 for문 과 같음
#     print(row[1].value)

# # 전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)
# for row in ws.iter_cols(): # 위에 for문 과 같음
#     print(row[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번쨰 열까지
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3) :
#     print(row[0].value, row[1].value) # 수학,영어

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3) :
    print(col[0].value, col[1].value)

wb.save("sample.xlsx")